import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Este programa carga un archivo Excel con calificaciones de alumnos,
# y agrega el formato de negrita a la columna "Nombre".

# Cargar el archivo Excel
archivo_excel = 'calificaciones_alumnos.xlsx'
df = pd.read_excel(archivo_excel)

# Guardar el archivo Excel modificado
archivo_modificado = 'calificaciones_alumnos_nombre_negrita.xlsx'
df.to_excel(archivo_modificado, index=False)

# Cargar el archivo Excel para modificarlo con openpyxl
wb = load_workbook(archivo_modificado)
ws = wb.active

# Definir el formato de negrita
negrita = Font(bold=True)

# Encontrar el índice de la columna "Nombre"
col_idx = None
for i, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
    if col[0].value == 'Nombre':
        col_idx = i
        break

# Aplicar el formato de negrita a las celdas de la columna "Nombre"
if col_idx:
    for cell in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        for c in cell:
            c.font = negrita

# Guardar el archivo Excel con el formato aplicado
wb.save(archivo_modificado)

print(f'El archivo modificado ha sido guardado como "{archivo_modificado}"')