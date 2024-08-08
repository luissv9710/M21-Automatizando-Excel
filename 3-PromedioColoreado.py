import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Este programa carga un archivo Excel con calificaciones de alumnos,
# colorea las celdas de la columna "Promedio" en verde si el valor es mayor o igual a 70,
# y en rojo en caso contrario.

# Cargar el archivo Excel
archivo_excel = 'calificaciones_alumnos.xlsx'
df = pd.read_excel(archivo_excel)

# Guardar el archivo Excel modificado
archivo_modificado = 'calificaciones_alumnos_promedio_color.xlsx'
df.to_excel(archivo_modificado, index=False)

# Colorear las celdas de la columna "Promedio" en verde si el valor es mayor o igual a 70, y en rojo en caso contrario
wb = load_workbook(archivo_modificado)
ws = wb.active

# Definir los colores
verde = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
rojo = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# Encontrar el índice de la columna "Promedio"
col_idx = None
for i, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
    if col[0].value == 'Promedio':
        col_idx = i
        break

# Aplicar los colores a las celdas de la columna "Promedio"
if col_idx:
    for cell in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        for c in cell:
            if c.value >= 70:
                c.fill = verde
            else:
                c.fill = rojo

# Guardar el archivo Excel con las celdas coloreadas
wb.save(archivo_modificado)

print(f'El archivo modificado ha sido guardado como "{archivo_modificado}"')